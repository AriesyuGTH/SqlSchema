from openpyxl import load_workbook

class ReadXls:
    def __init__(self,filename):
        self.wb = load_workbook(filename = filename) 

    def fieldTitle(self,sheetName,rowPoisition):
        """取得欄位名稱"""
        ws = self.wb[sheetName]
        rowObj = []
        for row in ws.iter_rows(min_row=rowPoisition, max_col=ws.max_column, max_row=rowPoisition): 
            rowObj = ([row[i].value for i in range(ws.max_column)])
        return rowObj

    def getSheetData(self,sheetName,startRow = 3):
        """取得指定資料表的內容"""
        ws = self.wb[sheetName]
        rows = []
        for row in ws.iter_rows(min_row=startRow, max_col=ws.max_column, max_row=ws.max_row): 
            rows.append([row[i] for i in range(ws.max_column)])
        return rows
        # for row in rows: 
        #     for cell in row:
        #         print(cell.value)

    def getSheetNames(self):
        return [sheetName for sheetName in self.wb.sheetnames if sheetName not in ['輸出摘要','Index']]
